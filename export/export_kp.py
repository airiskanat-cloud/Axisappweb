import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import os
import tempfile
from typing import Dict

def export_to_excel(order_data: dict, result_data: dict, output_dir: str = None) -> str:
    """
    Создает Excel файл с коммерческим предложением для окон/дверей
    Возвращает путь к созданному файлу
    
    Args:
        order_data: данные заказа
        result_data: результаты расчета
        output_dir: директория для сохранения
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Коммерческое предложение"
    
    # Стили
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=14, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Извлекаем данные
    common = order_data.get("common", {})
    order_num = common.get("order_number", "001")
    total_price = result_data.get("total_with_margin", 0)
    
    metrics = result_data.get("metrics", {})
    total_area = metrics.get("total_area", 0)
    
    # ШАПКА КОМПАНИИ
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Компания «AXIS»"
    cell.font = title_font
    cell.alignment = center_alignment
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Город Астана"
    cell.font = normal_font
    cell.alignment = center_alignment
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Тел.: +7 707 504 4040"
    cell.font = normal_font
    cell.alignment = center_alignment
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
    cell.font = title_font
    cell.alignment = center_alignment
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"Заказ № {order_num} от {datetime.datetime.now().strftime('%d.%m.%Y')}"
    cell.font = header_font
    cell.alignment = center_alignment
    
    row += 2
    
    # ИСПРАВЛЕНО: Упрощённое КП - только площадь и итого
    
    # ИТОГИ
    ws[f'A{row}'] = "Общая площадь:"
    ws[f'A{row}'].font = bold_font
    ws[f'E{row}'] = f"{total_area:.3f}"
    ws[f'E{row}'].font = normal_font
    ws[f'F{row}'] = "м²"
    ws[f'F{row}'].font = normal_font
    
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "ИТОГО к оплате:"
    cell.font = Font(name='Arial', size=14, bold=True)
    cell.fill = total_fill
    cell.alignment = left_alignment
    cell.border = thin_border
    
    ws[f'E{row}'] = f"{total_price:,.2f}"
    ws[f'E{row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'E{row}'].fill = total_fill
    ws[f'E{row}'].alignment = center_alignment
    ws[f'E{row}'].border = thin_border
    
    ws[f'F{row}'] = "₸"
    ws[f'F{row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'F{row}'].fill = total_fill
    ws[f'F{row}'].alignment = center_alignment
    ws[f'F{row}'].border = thin_border
    
    # Ширина столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    
    # Сохранение файла
    if output_dir is None:
        output_dir = tempfile.gettempdir()
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    filename = f"KP_AXIS_{order_num}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    return filepath


def export_facade_to_excel(facade_result: Dict, order_number: str = None, output_dir: str = None) -> str:
    """
    Экспорт результатов расчета фасадов в Excel
    
    Args:
        facade_result: результаты расчета фасада
        order_number: номер заказа
        output_dir: директория для сохранения
    
    Returns:
        Путь к созданному файлу
    """
    
    # ИСПРАВЛЕНО: Убрана проверка success - она опциональна
    if not facade_result:
        raise ValueError("Результаты расчета фасада отсутствуют")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "КП Фасад"
    
    # Стили
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=14, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # ШАПКА
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Компания «AXIS» - Фасадные системы"
    cell.font = title_font
    cell.alignment = center_alignment
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Город Астана"
    cell.font = normal_font
    cell.alignment = center_alignment
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Тел.: +7 707 504 4040"
    cell.font = normal_font
    cell.alignment = center_alignment
    
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
    cell.font = title_font
    cell.alignment = center_alignment
    
    row += 2
    
    # Извлекаем данные из результата
    facade_type = facade_result.get('facade_type', 'Фасад')
    total_area = facade_result.get('metrics', {}).get('total_area', 0)
    total_cost = facade_result.get('total_cost', 0)
    part3 = facade_result.get('part3_final', {})
    
    # Позиции - только общая площадь
    ws[f'A{row}'] = "Общая площадь:"
    ws[f'A{row}'].font = bold_font
    ws[f'E{row}'] = f"{total_area:.2f}"
    ws[f'E{row}'].font = normal_font
    ws[f'F{row}'] = "м²"
    ws[f'F{row}'].font = normal_font
    
    row += 2
    
    # ИТОГО (БЕЗ ДЕТАЛИЗАЦИИ)
    ws[f'A{row}'] = "ИТОГО к оплате:"
    ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'E{row}'] = f"{total_cost:,.2f}"
    ws[f'E{row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'F{row}'] = "₸"
    ws[f'F{row}'].font = Font(name='Arial', size=14, bold=True)
    
    # Ширина столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['E'].width = 20
    
    # Сохранение
    if output_dir is None:
        output_dir = tempfile.gettempdir()
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    if not order_number:
        order_number = f"FAC-{datetime.datetime.now().strftime('%Y%m%d%H%M')}"
    
    filename = f"KP_FACADE_{order_number}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    return filepath
