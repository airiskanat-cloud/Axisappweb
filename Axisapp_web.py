# =========================================
# Axis Pro GF — Calculator
# BASE CORE (NO BREAKING CHANGES)
# =========================================

import math
import ast
import operator as op
import json
import logging
import sys
import os
import datetime
import copy

import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials

# =========================================
# CONFIG
# =========================================

APP_TITLE = "Axis Pro GF — Фасад / Окна / Двери"

GSPREAD_SHEET_ID = "13kxXxhYNkMBhnltEZT6v2cdRu6aTF4_7wm7glqq45O8"

SHEET_REF1 = "СПРАВОЧНИК -1"
SHEET_REF2 = "СПРАВОЧНИК -2"
SHEET_REF3 = "СПРАВОЧНИК -3"
SHEET_USERS = "ПОЛЬЗОВАТЕЛИ"
SHEET_FORM = "ЗАПРОСЫ"

ENSURE_PERCENT = 0.65  # 65% обеспечение

SERVICE_ACCOUNT_PATH = "/etc/secrets/gcp_service_account.json"

# =========================================
# LOGGER
# =========================================

logger = logging.getLogger("axis_pro_gf")
if not logger.handlers:
    handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s"
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)

logger.setLevel(logging.INFO)

# =========================================
# UTILS
# =========================================

def normalize_key(value):
    """
    Используется ТОЛЬКО для поиска и диагностики.
    НЕ применяется для строгого сравнения справочников.
    """
    if value is None:
        return ""
    return " ".join(
        str(value)
        .replace("\xa0", " ")
        .replace(",", ".")
        .lower()
        .split()
    )


def safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        s = (
            str(value)
            .replace("\xa0", "")
            .replace(" ", "")
            .replace(",", ".")
        )
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def safe_int(value, default=0):
    try:
        return int(float(value))
    except Exception:
        return default


def get_field(row: dict, needle: str, default=None):
    """
    Поиск значения в строке Google Sheets
    по части названия колонки
    """
    needle = needle.lower()
    for k, v in row.items():
        if k and needle in str(k).lower():
            return v
    return default


def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# =========================================
# SAFE AST EVAL (FORMULA_PYTHON)
# =========================================

_ALLOWED_OPS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.Pow: op.pow,
    ast.USub: op.neg,
    ast.UAdd: op.pos,
    ast.Mod: op.mod,
}


def _eval_node(node, names):
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, names)

    if isinstance(node, ast.Constant):
        return node.value

    if isinstance(node, ast.Name):
        if node.id in names:
            return names[node.id]
        raise ValueError(f"Unknown variable: {node.id}")

    if isinstance(node, ast.BinOp):
        return _ALLOWED_OPS[type(node.op)](
            _eval_node(node.left, names),
            _eval_node(node.right, names),
        )

    if isinstance(node, ast.UnaryOp):
        return _ALLOWED_OPS[type(node.op)](
            _eval_node(node.operand, names)
        )

    if isinstance(node, ast.Call):
        if isinstance(node.func, ast.Attribute):
            if (
                isinstance(node.func.value, ast.Name)
                and node.func.value.id == "math"
            ):
                fn = getattr(math, node.func.attr)
                args = [_eval_node(a, names) for a in node.args]
                return fn(*args)

        if isinstance(node.func, ast.Name):
            if node.func.id in ("min", "max", "round"):
                args = [_eval_node(a, names) for a in node.args]
                return globals()[node.func.id](*args)

    raise ValueError("Unsafe expression detected")


def safe_eval(formula: str, context: dict) -> float:
    """
    Безопасное вычисление формул из СПРАВОЧНИК-1
    """
    if not formula:
        return 0.0
    try:
        ctx = {k: safe_float(v) for k, v in context.items()}
        ctx["math"] = math
        node = ast.parse(str(formula), mode="eval")
        return float(_eval_node(node, ctx))
    except Exception as e:
        logger.error(f"FORMULA ERROR [{formula}] -> {e}")
        return 0.0

# =========================================
# GOOGLE SHEETS CLIENT
# =========================================

cclass GoogleSheets:

    @st.cache_resource
    def auth(_self):
        if not os.path.exists(SERVICE_ACCOUNT_PATH):
            st.error("❌ Не найден файл сервисного аккаунта Google")
            st.stop()

        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_PATH,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        return gspread.authorize(creds)

    def __init__(self, sheet_id: str):
        self.client = self.auth()
        self.book = self.client.open_by_key(sheet_id)
        self._cache = {}

    def ws(self, name: str):
        if name not in self._cache:
            self._cache[name] = self.book.worksheet(name)
        return self._cache[name]

    @st.cache_data(ttl=1800)
    def read(self, sheet_name: str):
        ws = self.ws(sheet_name)
        return ws.get_all_records()

    def append(self, sheet_name: str, row: list):
        ws = self.ws(sheet_name)
        ws.append_row(row, value_input_option="USER_ENTERED")

# =========================================
# LOGIN
# =========================================

def login(gs: GoogleSheets) -> bool:
    if "user" in st.session_state:
        return True

    st.sidebar.title("🔐 Вход")

    login_value = st.sidebar.text_input("Логин")
    password_value = st.sidebar.text_input(
        "Пароль",
        type="password",
    )

    if st.sidebar.button("Войти"):
        users = gs.read(SHEET_USERS)
        for u in users:
            login_db = str(get_field(u, "логин", "")).strip()
            password_db = str(get_field(u, "пароль", "")).strip()

            if login_db == login_value and password_db == password_value:
                st.session_state["user"] = login_db
                st.session_state["login_time"] = now_str()
                st.rerun()

        st.sidebar.error("Неверный логин или пароль")

    return False
# =========================================
# POSITIONS & GEOMETRY
# =========================================

def build_sash_context(sash: dict) -> dict:
    """
    Контекст одной створки.
    Используется для расчётов стекла, уплотнителей и пр.
    """
    sw = safe_float(sash.get("sash_width_mm", 0))
    sh = safe_float(sash.get("sash_height_mm", 0))
    qty = safe_int(sash.get("qty", 1), 1)

    area = (sw * sh) / 1_000_000
    perimeter = 2 * (sw + sh) / 1000

    return {
        "sash_w": sw,
        "sash_h": sh,
        "glass_w": sw,
        "glass_h": sh,
        "sash_area": area,
        "sash_perimeter": perimeter,
        "qty": qty,
    }


def build_position_geometry(position: dict) -> dict:
    """
    Геометрия ОДНОЙ позиции (изделия).
    Совпадает по смыслу с листом 'Расчет по габаритам'.
    """

    W = safe_float(position.get("width_mm", 0))
    H = safe_float(position.get("height_mm", 0))
    qty = safe_int(position.get("qty", 1), 1)

    left = safe_float(position.get("left_mm", 0))
    center = safe_float(position.get("center_mm", 0))
    right = safe_float(position.get("right_mm", 0))
    top = safe_float(position.get("top_mm", 0))

    # Базовые величины
    area = (W * H) / 1_000_000
    perimeter = 2 * (W + H) / 1000

    # Импосты
    n_vert = sum(1 for v in (left, center, right) if v > 0)
    n_imp_vert = max(0, n_vert)
    n_imp_hor = 1 if top > 0 else 0
    n_impost = n_imp_vert + n_imp_hor

    # Прямоугольники (как в Excel)
    n_frame_rect = 1 + n_impost
    n_corners = n_frame_rect * 4

    # Створки
    sashes = position.get("sashes", [])
    sash_contexts = [build_sash_context(s) for s in sashes]

    sash_area_total = sum(s["sash_area"] * s["qty"] for s in sash_contexts)
    sash_perimeter_total = sum(
        s["sash_perimeter"] * s["qty"] for s in sash_contexts
    )

    return {
        # Габариты изделия
        "W": W,
        "H": H,
        "width": W,
        "height": H,
        "area": area,
        "perimeter": perimeter,
        "qty": qty,

        # Импосты
        "left": left,
        "center": center,
        "right": right,
        "top": top,
        "n_imp_vert": n_imp_vert,
        "n_imp_hor": n_imp_hor,
        "n_impost": n_impost,

        # Конструктив
        "n_frame_rect": n_frame_rect,
        "n_corners": n_corners,

        # Створки
        "n_sash": len(sash_contexts),
        "sashes": sash_contexts,
        "sash_area_total": sash_area_total,
        "sash_perimeter_total": sash_perimeter_total,

        # Для формул
        "glass_w": W,
        "glass_h": H,
    }


def build_positions_aggregate(positions: list) -> dict:
    """
    Агрегация ВСЕХ позиций заказа.
    Это ИСТОЧНИК ИСТИНЫ для итогов.
    """

    total_area = 0.0
    total_perimeter = 0.0
    total_sash_area = 0.0
    total_sash_perimeter = 0.0
    total_qty = 0

    for p in positions:
        geo = p["geometry"]

        total_area += geo["area"] * geo["qty"]
        total_perimeter += geo["perimeter"] * geo["qty"]
        total_sash_area += geo["sash_area_total"] * geo["qty"]
        total_sash_perimeter += geo["sash_perimeter_total"] * geo["qty"]
        total_qty += geo["qty"]

    return {
        "total_area": total_area,
        "total_perimeter": total_perimeter,
        "total_sash_area": total_sash_area,
        "total_sash_perimeter": total_sash_perimeter,
        "total_positions_qty": total_qty,
    }

# =========================================
# POSITION FACTORY
# =========================================

def create_position(
    product_type: str,
    profile_system: str,
    width_mm: float,
    height_mm: float,
    qty: int,
    left_mm: float = 0,
    center_mm: float = 0,
    right_mm: float = 0,
    top_mm: float = 0,
    sashes: list = None,
) -> dict:
    """
    Унифицированное создание позиции.
    """
    if sashes is None:
        sashes = []

    position = {
        "product_type": product_type,
        "profile_system": profile_system,
        "width_mm": width_mm,
        "height_mm": height_mm,
        "qty": qty,
        "left_mm": left_mm,
        "center_mm": center_mm,
        "right_mm": right_mm,
        "top_mm": top_mm,
        "sashes": sashes,
    }

    position["geometry"] = build_position_geometry(position)
    return position

# =========================================
# DEBUG HELPERS (OPTIONAL UI)
# =========================================

def debug_positions(positions: list):
    rows = []
    for i, p in enumerate(positions, 1):
        g = p["geometry"]
        rows.append({
            "№": i,
            "Тип изделия": p["product_type"],
            "Система": p["profile_system"],
            "Ширина мм": g["W"],
            "Высота мм": g["H"],
            "Кол-во": g["qty"],
            "Площадь м²": round(g["area"], 3),
            "Периметр м": round(g["perimeter"], 3),
            "Створок": g["n_sash"],
            "Площадь створок м²": round(g["sash_area_total"], 3),
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True)
# =========================================
# FORMULA CONTEXT BUILDER (СПРАВОЧНИК-1)
# =========================================

def build_formula_context(position: dict) -> dict:
    """
    Формирует КОНТЕКСТ для формул из СПРАВОЧНИК-1.
    Логика повторяет Excel:
    - count
    - W / H
    - glass_w / glass_h
    - створки
    - импосты
    """

    geo = position["geometry"]

    # --- БАЗОВЫЕ ---
    context = {
        # как в Excel
        "count": geo["qty"],

        # габариты изделия
        "W": geo["W"],
        "H": geo["H"],
        "width": geo["W"],
        "height": geo["H"],

        # площадь / периметр
        "area": geo["area"],
        "perimeter": geo["perimeter"],

        # импосты
        "left": geo["left"],
        "center": geo["center"],
        "right": geo["right"],
        "top": geo["top"],
        "n_imp_vert": geo["n_imp_vert"],
        "n_imp_hor": geo["n_imp_hor"],
        "n_impost": geo["n_impost"],

        # конструктив
        "n_frame_rect": geo["n_frame_rect"],
        "n_corners": geo["n_corners"],

        # створки
        "n_sash": geo["n_sash"],
        "sash_area_total": geo["sash_area_total"],
        "sash_perimeter_total": geo["sash_perimeter_total"],
    }

    # --- GLASS SIZE ---
    # По Excel: если есть створки — стекло по створкам,
    # если нет — по изделию
    if geo["n_sash"] > 0 and geo["sashes"]:
        # Берём первую створку как базу (как в v15)
        first_sash = geo["sashes"][0]
        context["glass_w"] = first_sash["glass_w"]
        context["glass_h"] = first_sash["glass_h"]
    else:
        context["glass_w"] = geo["W"]
        context["glass_h"] = geo["H"]

    # --- ДОПОЛНИТЕЛЬНО ДЛЯ СЛОЖНЫХ ФОРМУЛ ---
    context["total_sash_area"] = geo["sash_area_total"]
    context["total_sash_perimeter"] = geo["sash_perimeter_total"]

    return context


def build_all_formula_contexts(positions: list) -> list:
    """
    Возвращает список контекстов — ПО ОДНОМУ НА КАЖДУЮ ПОЗИЦИЮ.
    MaterialCalculator будет проходить по ним.
    """
    contexts = []
    for p in positions:
        ctx = build_formula_context(p)
        contexts.append({
            "product_type": p["product_type"],
            "profile_system": p["profile_system"],
            "context": ctx,
        })
    return contexts


# =========================================
# DEBUG: ПРОВЕРКА ФОРМУЛЬНОГО КОНТЕКСТА
# =========================================

def debug_formula_contexts(positions: list):
    """
    Отладочный вывод контекста формул
    — помогает понять, ПОЧЕМУ формула дала 0
    """
    rows = []
    for i, p in enumerate(positions, 1):
        ctx = build_formula_context(p)
        rows.append({
            "№": i,
            "Тип изделия": p["product_type"],
            "Система": p["profile_system"],
            "count": ctx["count"],
            "W": ctx["W"],
            "H": ctx["H"],
            "glass_w": ctx["glass_w"],
            "glass_h": ctx["glass_h"],
            "n_sash": ctx["n_sash"],
            "area": round(ctx["area"], 3),
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True)
# =========================================
# MATERIAL CALCULATOR (СПРАВОЧНИК-1)
# =========================================

class MaterialCalculator:
    """
    Полный расчёт материалов.
    Логика соответствует Excel:
    - формулы из СПРАВОЧНИК-1
    - контекст из ЧАСТИ 3
    - строгая привязка:
        Тип изделия
        Система профиля
    """

    def __init__(self, gs: GoogleSheets):
        self.gs = gs

    def calculate(self, positions: list):
        """
        positions — список позиций (из create_position)
        """
        ref1 = self.gs.read(SHEET_REF1)

        results = []
        total_sum = 0.0

        # Контексты формул по позициям
        formula_contexts = build_all_formula_contexts(positions)

        for row in ref1:
            row_product_type = str(
                get_field(row, "тип издел", "")
            ).strip()

            row_profile_system = str(
                get_field(row, "система проф", "")
            ).strip()

            element_type = str(
                get_field(row, "тип элемент", "")
            ).strip()

            product_name = str(
                get_field(row, "товар", "")
            ).strip()

            formula = get_field(row, "формула_python")
            if not formula:
                continue

            price = safe_float(
                get_field(row, "цена за", 0)
            )

            norm = safe_float(
                get_field(row, "кол-во норм", 1),
                1,
            )

            qty_total = 0.0

            # === ПРОХОД ПО ПОЗИЦИЯМ ===
            for ctx_item in formula_contexts:
                if row_product_type and row_product_type != ctx_item["product_type"]:
                    continue

                if row_profile_system and row_profile_system != ctx_item["profile_system"]:
                    continue

                ctx = ctx_item["context"]

                # === РАСЧЁТ ПО ФОРМУЛЕ ===
                value = safe_eval(formula, ctx)

                # === ТИП ЭЛЕМЕНТА ===
                # (логика как в Excel)
                if normalize_key(element_type) in (
                    "на изделие",
                    "на позицию",
                ):
                    qty_total += value

                elif normalize_key(element_type) in (
                    "на створку",
                ):
                    qty_total += value * ctx.get("n_sash", 1)

                elif normalize_key(element_type) in (
                    "на м2",
                    "на м²",
                ):
                    qty_total += value * ctx.get("area", 0)

                else:
                    # по умолчанию — как есть
                    qty_total += value

            if qty_total <= 0:
                continue

            # === НОРМИРОВАНИЕ (ХЛЫСТЫ, КОМПЛЕКТЫ) ===
            if norm > 0:
                ship_qty = math.ceil(qty_total / norm)
                real_qty = ship_qty * norm
            else:
                real_qty = qty_total

            sum_row = real_qty * price
            total_sum += sum_row

            results.append({
                "Тип изделия": row_product_type,
                "Система профиля": row_profile_system,
                "Тип элемента": element_type,
                "Товар": product_name,
                "Формула": formula,
                "Факт. расход": round(qty_total, 3),
                "К отгрузке": round(real_qty, 3),
                "Цена": price,
                "Сумма": round(sum_row, 2),
            })

        return results, round(total_sum, 2)
# =========================================
# REF2: GLASS & SERVICES (СПРАВОЧНИК-2)
# =========================================

def parse_ref2_table(ref2_rows: list) -> dict:
    """
    Преобразует СПРАВОЧНИК-2 в словарь:
    {
        "двойной": {
            "glass_price": 9000,
            "panel_name": "Ламбри без термо",
            "panel_price": 2248,
            "toning_name": "Есть",
            "toning_price": 2000,
            "assembly_name": "Есть",
            "assembly_price": 10000,
            "montage_name": "Монтаж",
            "montage_price": 10000,
        },
        ...
    }
    """
    table = {}

    for row in ref2_rows:
        glass_type = str(
            get_field(row, "Тип стеклопакета", "")
        ).strip()

        if not glass_type:
            continue

        table[glass_type] = {
            "glass_price": safe_float(
                get_field(row, "Стоимость стеклопакета", 0)
            ),
            "panel_name": str(
                get_field(row, "Панели", "")
            ).strip(),
            "panel_price": safe_float(
                get_field(row, "Стоимость Панелей", 0)
            ),
            "toning_name": str(
                get_field(row, "Тонировка", "")
            ).strip(),
            "toning_price": safe_float(
                get_field(row, "Стоимость тонировки", 0)
            ),
            "assembly_name": str(
                get_field(row, "Сборка", "")
            ).strip(),
            "assembly_price": safe_float(
                get_field(row, "Стоимость сборки", 0)
            ),
            "montage_name": str(
                get_field(row, "Монтаж", "")
            ).strip(),
            "montage_price": safe_float(
                get_field(row, "Стоимость монтаж", 0)
            ),
        }

    return table


# =========================================
# GLASS & SERVICES CALCULATOR
# =========================================

class GlassServiceCalculator:
    """
    Расчёт стеклопакета и услуг
    строго по СПРАВОЧНИК-2
    """

    def __init__(self, gs: GoogleSheets):
        self.ref2 = parse_ref2_table(gs.read(SHEET_REF2))

    def get_glass_types(self) -> list:
        return list(self.ref2.keys())

    def calculate(self, glass_type: str, aggregates: dict):
        """
        aggregates — результат build_positions_aggregate
        """
        if glass_type not in self.ref2:
            return [], 0.0

        data = self.ref2[glass_type]

        rows = []
        total = 0.0

        area = aggregates["total_area"]

        # --- Стеклопакет ---
        glass_sum = area * data["glass_price"]
        rows.append((
            f"Стеклопакет ({glass_type})",
            data["glass_price"],
            "м²",
            glass_sum,
        ))
        total += glass_sum

        # --- Панели ---
        if data["panel_price"] > 0:
            panel_sum = area * data["panel_price"]
            rows.append((
                f"Панели ({data['panel_name']})",
                data["panel_price"],
                "м²",
                panel_sum,
            ))
            total += panel_sum

        # --- Тонировка ---
        if data["toning_price"] > 0:
            toning_sum = area * data["toning_price"]
            rows.append((
                f"Тонировка ({data['toning_name']})",
                data["toning_price"],
                "м²",
                toning_sum,
            ))
            total += toning_sum

        # --- Сборка ---
        if data["assembly_price"] > 0:
            assembly_sum = area * data["assembly_price"]
            rows.append((
                f"Сборка ({data['assembly_name']})",
                data["assembly_price"],
                "м²",
                assembly_sum,
            ))
            total += assembly_sum

        # --- Монтаж ---
        if data["montage_price"] > 0:
            montage_sum = area * data["montage_price"]
            rows.append((
                f"Монтаж ({data['montage_name']})",
                data["montage_price"],
                "м²",
                montage_sum,
            ))
            total += montage_sum

        return rows, round(total, 2)


# =========================================
# DEBUG: REF2 VIEWER
# =========================================

def debug_ref2(gs: GoogleSheets):
    st.subheader("СПРАВОЧНИК-2 (отладка)")
    df = pd.DataFrame(gs.read(SHEET_REF2))
    st.dataframe(df, use_container_width=True)
# =========================================
# FACADE ENGINEERING LOGIC
# =========================================

# Таблица стоек фасада (из Excel "РАСЧЕТ ВЕТРОВОЙ НАГРУЗКИ")
# Упорядочена по возрастанию Jx
FACADE_STANDS = [
    {"profile": "90-5035",  "Jx": 79},
    {"profile": "100-5009", "Jx": 117},
    {"profile": "110-5034", "Jx": 126},
    {"profile": "130-5033", "Jx": 190},
    {"profile": "150-5032", "Jx": 277},
    {"profile": "170-5010", "Jx": 403},
    {"profile": "160-5005", "Jx": 422},
    {"profile": "200-5006", "Jx": 851},
]


def calc_required_jx(height_mm: float) -> float:
    """
    Расчёт требуемого момента инерции Jx.
    Основан на концепте Excel:
    зависимость квадратичная от высоты фасада.
    """
    H = safe_float(height_mm) / 1000.0  # м
    # коэффициент взят из концепта Excel (без фантазий)
    required_jx = 55 * (H ** 2)
    return required_jx


def select_facade_stand(height_mm: float) -> dict:
    """
    Выбор стойки фасада:
    берём первую стойку, у которой Jx >= требуемого
    """
    req_jx = calc_required_jx(height_mm)

    for stand in FACADE_STANDS:
        if stand["Jx"] >= req_jx:
            return {
                "required_jx": req_jx,
                "selected_profile": stand["profile"],
                "selected_jx": stand["Jx"],
            }

    # если ни одна не подошла — берём максимальную
    last = FACADE_STANDS[-1]
    return {
        "required_jx": req_jx,
        "selected_profile": last["profile"],
        "selected_jx": last["Jx"],
    }


# =========================================
# FACADE POSITION EXTENSION
# =========================================

def apply_facade_engineering(position: dict) -> dict:
    """
    Добавляет инженерные данные фасада в позицию.
    Вызывается ТОЛЬКО для product_type == 'Фасад'
    """
    geo = position["geometry"]

    height_mm = geo["H"]

    stand_step = safe_float(
        position.get("stand_step_mm", 1200)
    )

    stand_info = select_facade_stand(height_mm)

    position["facade_engineering"] = {
        "height_mm": height_mm,
        "stand_step_mm": stand_step,
        "required_jx": stand_info["required_jx"],
        "selected_stand_profile": stand_info["selected_profile"],
        "selected_stand_jx": stand_info["selected_jx"],
    }

    return position


# =========================================
# FACADE HELPERS
# =========================================

def process_facade_positions(positions: list) -> list:
    """
    Проходит по всем позициям,
    применяет инженерную логику ТОЛЬКО к фасаду
    """
    processed = []

    for p in positions:
        if p["product_type"] == "Фасад":
            processed.append(apply_facade_engineering(p))
        else:
            processed.append(p)

    return processed


# =========================================
# DEBUG: FACADE ENGINEERING
# =========================================

def debug_facade(positions: list):
    rows = []

    for i, p in enumerate(positions, 1):
        if p["product_type"] != "Фасад":
            continue

        fe = p.get("facade_engineering", {})
        rows.append({
            "№": i,
            "Высота фасада мм": fe.get("height_mm"),
            "Шаг стоек мм": fe.get("stand_step_mm"),
            "Требуемый Jx": round(fe.get("required_jx", 0), 2),
            "Выбранная стойка": fe.get("selected_stand_profile"),
            "Jx стойки": fe.get("selected_stand_jx"),
        })

    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True)
    else:
        st.info("Фасадных позиций нет")
# =========================================
# REQUESTS / HISTORY
# =========================================

def serialize_positions(positions: list) -> str:
    """
    Сериализация позиций в JSON
    (для сохранения истории и восстановления расчёта)
    """
    clean_positions = []

    for p in positions:
        cp = copy.deepcopy(p)

        # geometry и инженерия тоже сохраняем
        # (это важно для воспроизводимости)
        clean_positions.append(cp)

    return json.dumps(clean_positions, ensure_ascii=False)


def serialize_materials(material_rows: list) -> str:
    """
    Сериализация материалов
    """
    return json.dumps(material_rows, ensure_ascii=False)


def save_request(
    gs: GoogleSheets,
    user: str,
    positions: list,
    aggregates: dict,
    material_sum: float,
    glass_rows: list,
    glass_sum: float,
    total_sum: float,
):
    """
    Сохраняет запрос в лист ЗАПРОСЫ.
    Каждая строка = один расчёт.
    """

    row = [
        now_str(),                      # Дата
        user,                           # Пользователь
        len(positions),                # Кол-во позиций
        round(aggregates["total_area"], 3),
        round(aggregates["total_perimeter"], 3),
        round(material_sum, 2),         # Материалы
        round(glass_sum, 2),            # Стекло + услуги
        round(material_sum + glass_sum, 2),  # База
        round((material_sum + glass_sum) * ENSURE_PERCENT, 2),  # 65%
        round(total_sum, 2),            # ИТОГО
        serialize_positions(positions),
        serialize_materials(glass_rows),
    ]

    gs.append(SHEET_FORM, row)


# =========================================
# HISTORY VIEWER
# =========================================

def show_history(gs: GoogleSheets):
    """
    Отображает историю запросов
    """
    st.subheader("📜 История запросов")

    rows = gs.read(SHEET_FORM)
    if not rows:
        st.info("История пуста")
        return

    df = pd.DataFrame(rows)

    st.dataframe(df, use_container_width=True)


# =========================================
# DEBUG: REQUEST PREVIEW
# =========================================

def debug_request_preview(
    positions: list,
    aggregates: dict,
    material_sum: float,
    glass_sum: float,
    total_sum: float,
):
    st.subheader("Отладка запроса")

    st.write("Позиций:", len(positions))
    st.write("Площадь:", aggregates["total_area"])
    st.write("Периметр:", aggregates["total_perimeter"])
    st.write("Материалы:", material_sum)
    st.write("Стекло + услуги:", glass_sum)
    st.write("ИТОГО:", total_sum)
# =========================================
# COMMERCIAL OFFER (KP)
# =========================================

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile


def generate_kp_excel(
    template_path: str,
    user: str,
    aggregates: dict,
    material_sum: float,
    glass_rows: list,
    glass_sum: float,
    total_sum: float,
):
    """
    Генерация коммерческого предложения
    на основе Excel-шаблона v15
    """

    wb = load_workbook(template_path)
    ws = wb.active

    # === ШАПКА ===
    ws["B2"] = f"Коммерческое предложение"
    ws["B3"] = f"Менеджер: {user}"
    ws["B4"] = f"Дата: {now_str()}"

    # === ГАБАРИТЫ ===
    ws["B6"] = aggregates["total_area"]
    ws["B7"] = aggregates["total_perimeter"]

    # === ТАБЛИЦА РАСЧЁТА ===
    start_row = 10
    r = start_row

    # Материалы
    ws[f"A{r}"] = "Материалы"
    ws[f"D{r}"] = material_sum
    r += 1

    # Стекло и услуги
    for row in glass_rows:
        ws[f"A{r}"] = row[0]
        ws[f"B{r}"] = row[2]
        ws[f"C{r}"] = row[1]
        ws[f"D{r}"] = row[3]
        r += 1

    # Итоги
    ws[f"A{r}"] = "Итого"
    ws[f"D{r}"] = material_sum + glass_sum
    r += 1

    ws[f"A{r}"] = "Обеспечение 65%"
    ws[f"D{r}"] = (material_sum + glass_sum) * ENSURE_PERCENT
    r += 1

    ws[f"A{r}"] = "ИТОГО К ОПЛАТЕ"
    ws[f"D{r}"] = total_sum

    # === СОХРАНЕНИЕ ===
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return tmp.name


# =========================================
# MAIN APPLICATION
# =========================================

def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title("🏗️ Axis Pro GF — Калькулятор")

    gs = GoogleSheets(GSPREAD_SHEET_ID)
    if not login(gs):
        st.stop()

    # ---------- SIDEBAR ----------
    with st.sidebar:
        st.header("Параметры заказа")

        product_type = st.selectbox(
            "Тип изделия",
            [
                "Окно с откр.",
                "Окно  глух.",
                "Дверь 2-х створч.",
                "Дверь 1 створч.",
                "Фасад",
            ],
        )

        profile_system = st.selectbox(
            "Система профиля",
            [
                "ALG 2030-63C",
                "ALG 2030-55C",
                "ALG 2030-73C",
                "ALG 2030-45C",
                "ALG 2030-Slim",
                "Ruit 50F",
            ],
        )

    # ---------- ПОЗИЦИИ ----------
    st.header("📦 Позиции")

    positions = []

    positions_count = st.number_input(
        "Количество позиций",
        min_value=1,
        step=1,
        value=1,
    )

    for i in range(int(positions_count)):
        st.subheader(f"Позиция #{i+1}")

        c1, c2, c3 = st.columns(3)
        w = c1.number_input("Ширина, мм", key=f"w_{i}", step=10.0)
        h = c2.number_input("Высота, мм", key=f"h_{i}", step=10.0)
        qty = c3.number_input("Кол-во", key=f"q_{i}", min_value=1, step=1)

        positions.append(
            create_position(
                product_type=product_type,
                profile_system=profile_system,
                width_mm=w,
                height_mm=h,
                qty=qty,
            )
        )

    # ---------- АГРЕГАЦИЯ ----------
    aggregates = build_positions_aggregate(positions)

    # ---------- КАЛЬКУЛЯТОРЫ ----------
    mat_calc = MaterialCalculator(gs)
    glass_calc = GlassServiceCalculator(gs)

    glass_type = st.selectbox(
        "Тип стеклопакета",
        glass_calc.get_glass_types(),
    )

    if st.button("🚀 Рассчитать", type="primary"):
        # Материалы
        mat_rows, mat_sum = mat_calc.calculate(positions)

        # Стекло и услуги
        glass_rows, glass_sum = glass_calc.calculate(
            glass_type,
            aggregates,
        )

        base_sum = mat_sum + glass_sum
        ensure = base_sum * ENSURE_PERCENT
        total = base_sum + ensure

        # ---------- СОХРАНЕНИЕ ЗАПРОСА ----------
        save_request(
            gs=gs,
            user=st.session_state.get("user", ""),
            positions=positions,
            aggregates=aggregates,
            material_sum=mat_sum,
            glass_rows=glass_rows,
            glass_sum=glass_sum,
            total_sum=total,
        )

        # ---------- ВЫВОД ----------
        st.success(f"ИТОГО К ОПЛАТЕ: {round(total, 2)}")

        tab1, tab2, tab3 = st.tabs(
            ["📐 Геометрия", "🧱 Материалы", "💰 Итог"]
        )

        with tab1:
            st.metric("Общая площадь, м²", round(aggregates["total_area"], 3))
            st.metric("Общий периметр, м", round(aggregates["total_perimeter"], 3))

        with tab2:
            st.dataframe(pd.DataFrame(mat_rows), use_container_width=True)
            st.write(f"Материалы: {mat_sum}")

        with tab3:
            st.dataframe(
                pd.DataFrame(
                    glass_rows,
                    columns=["Наименование", "Цена", "Ед.", "Сумма"],
                ),
                use_container_width=True,
            )
            st.write(f"База: {base_sum}")
            st.write(f"Обеспечение 65%: {ensure}")
            st.write(f"ИТОГО: {total}")

        # ---------- КП ----------
        st.subheader("📄 Коммерческое предложение")

        kp_template = st.file_uploader(
            "Загрузите шаблон КП (v15)",
            type=["xlsx"],
        )

        if kp_template:
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
                f.write(kp_template.read())
                template_path = f.name

            kp_path = generate_kp_excel(
                template_path=template_path,
                user=st.session_state.get("user", ""),
                aggregates=aggregates,
                material_sum=mat_sum,
                glass_rows=glass_rows,
                glass_sum=glass_sum,
                total_sum=total,
            )

            with open(kp_path, "rb") as f:
                st.download_button(
                    "⬇️ Скачать КП",
                    f,
                    file_name="Коммерческое_предложение.xlsx",
                )

    # ---------- ИСТОРИЯ ----------
    st.markdown("---")
    show_history(gs)


# =========================================
# ENTRY POINT
# =========================================

if __name__ == "__main__":
    main()
