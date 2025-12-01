import math
import os
import sys
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# ======================================
# ПУТИ К ФАЙЛАМ
# ======================================

def resource_path(relative_path: str) -> str:
    """
    Возвращает корректный путь к файлу как при обычном запуске,
    так и в упакованном PyInstaller-приложении.
    Для веб-версии по сути просто путь относительно файла.
    """
    if hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)


TEMPLATE_EXCEL_NAME = "Справочник для Axis Pro GF.xlsx"
EXCEL_FILE = resource_path(TEMPLATE_EXCEL_NAME)

SHEET_REF1 = "СПРАВОЧНИК -1"
SHEET_REF2 = "СПРАВОЧНИК -2"
SHEET_REF3 = "СПРАВОЧНИК -3"
SHEET_FORM = "ЗАПРОСЫ"
SHEET_GABARITS = "Расчет по габаритам"
SHEET_MATERIAL = "Расчетом расходов материалов"
SHEET_FINAL = "Итоговый расчет с монтажом"

FORM_HEADER = [
    "Номер заказа", "№ позиции",
    "Тип изделия", "Вид изделия", "Створки",
    "Профильная система",
    "Толщина стеклопакета",
    "Тип стеклопакета",
    "Заполнение",
    "Ширина, мм", "Высота, мм",
    "LEFT, мм", "CENTER, мм", "RIGHT, мм", "TOP, мм",
    "Ширина створки, мм", "Высота створки, мм",
    "Кол-во Nwin",
    "Нарезка", "Тонировка", "Сборка", "Монтаж"
]


# ======================================
# РАБОТА С EXCEL
# ======================================

class ExcelClient:
    def __init__(self, filename: str):
        self.filename = filename
        if not os.path.exists(self.filename):
            wb = Workbook()
            wb.save(self.filename)
        self.load()

    def load(self):
        self.wb = load_workbook(self.filename, data_only=True)

    def save(self):
        self.wb.save(self.filename)

    def ws(self, name: str):
        if name in self.wb.sheetnames:
            return self.wb[name]
        ws = self.wb.create_sheet(name)
        self.save()
        return ws

    def read_records(self, sheet_name: str):
        ws = self.ws(sheet_name)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = rows[0]
        data_rows = rows[1:]
        records = []
        for r in data_rows:
            if all(v is None for v in r):
                continue
            rec = {}
            for i, key in enumerate(header):
                if key is None:
                    continue
                rec[str(key)] = r[i] if i < len(r) else None
            records.append(rec)
        return records

    def clear_and_write(self, sheet_name: str, header: list, rows: list):
        ws = self.ws(sheet_name)
        ws.delete_rows(1, ws.max_row or 1)
        if header:
            ws.append(header)
        for row in rows:
            ws.append(row)
        self.save()

    def append_form_row(self, row: list):
        ws = self.ws(SHEET_FORM)
        if ws.max_row == 1 and all(c.value is None for c in ws[1]):
            ws.append(FORM_HEADER)
        ws.append(row)
        self.save()


# ======================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ======================================

def safe_float(value, default=0.0):
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return default


def safe_int(value, default=0):
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return default


def get_field(row: dict, needle: str, default=None):
    needle = needle.lower()
    for k in row.keys():
        if k is None:
            continue
        if needle in str(k).lower():
            return row[k]
    return default


def eval_formula(formula: str, context: dict) -> float:
    """
    Считает формулу на Python для ОДНОЙ позиции.
    """
    formula = (formula or "").strip()
    if not formula:
        return 0.0

    allowed_names = {
        "width": context.get("width", 0.0),
        "height": context.get("height", 0.0),
        "left": context.get("left", 0.0),
        "center": context.get("center", 0.0),
        "right": context.get("right", 0.0),
        "top": context.get("top", 0.0),
        "sash_width": context.get("sash_width", 0.0),
        "sash_height": context.get("sash_height", 0.0),
        "area": context.get("area", 0.0),
        "perimeter": context.get("perimeter", 0.0),
        "qty": context.get("qty", 0.0),
        "nsash": context.get("nsash", 1),
        "n_sash_active": context.get("n_sash_active", 1),
        "n_sash_passive": context.get("n_sash_passive", 0),
        "hinges_per_sash": context.get("hinges_per_sash", 3),
        "n_rect": context.get("n_rect", 1),
        "n_impost": context.get("n_impost", 0),
        "N_impost": context.get("n_impost", 0),
        "math": math,
        "max": max,
        "min": min,
    }

    allowed_names.update({
        "n_imp_vert": context.get("n_imp_vert", 0),
        "n_imp_hor": context.get("n_imp_hor", 0),
        "n_frame_rect": context.get("n_frame_rect", 1),
        "n_corners": context.get("n_corners", 4),
        "n_nodes_12": context.get("n_nodes_12", 0),
        "n_nodes_19": context.get("n_nodes_19", 0),
        "n_nodes_6_5": context.get("n_nodes_6_5", 0),
        "n_nodes_17_2": context.get("n_nodes_17_2", 0),
        "n_nodes_42": context.get("n_nodes_42", 0),
        "Nwin": context.get("qty", 0.0),
    })

    try:
        result = eval(formula, {"__builtins__": {}}, allowed_names)
        return float(result)
    except Exception as e:
        print(f"Ошибка в формуле '{formula}': {e}")
        return 0.0


# ======================================
# РАСЧЕТ ПО ГАБАРИТАМ (СПРАВОЧНИК -3)
# ======================================

class GabaritCalculator:
    HEADER = ["Тип элемента", "Фактическое значение"]

    def __init__(self, excel_client: ExcelClient):
        self.excel = excel_client

    def calculate(self, order: dict, positions: list):
        ref_rows = self.excel.read_records(SHEET_REF3)
        if not ref_rows:
            return [], 0.0

        try:
            nsash = int(order.get("sashes", "1"))
        except ValueError:
            nsash = 1
        n_sash_active = 1 if nsash >= 1 else 0
        n_sash_passive = max(nsash - 1, 0)
        hinges_per_sash = 3

        total_area = sum(p["area_m2"] * p["Nwin"] for p in positions)

        gabarit_values = []

        for row in ref_rows:
            type_elem = get_field(row, "тип элемент", "")
            formula = get_field(row, "формула_python", "")
            if not type_elem or not formula:
                continue

            total_value = 0.0

            for p in positions:
                width = p["width_mm"]
                height = p["height_mm"]
                left = p["left_mm"]
                center = p["center_mm"]
                right = p["right_mm"]
                top = p["top_mm"]
                sash_w = p["sash_width_mm"]
                sash_h = p["sash_height_mm"]
                area = p["area_m2"]
                perimeter = p["perimeter_m"]
                qty = p["Nwin"]

                n_rect = 1 + (1 if left > 0 else 0) + (1 if top > 0 else 0)
                n_impost = 0

                ctx = {
                    "width": width,
                    "height": height,
                    "left": left,
                    "center": center,
                    "right": right,
                    "top": top,
                    "sash_width": sash_w,
                    "sash_height": sash_h,
                    "area": area,
                    "perimeter": perimeter,
                    "qty": qty,
                    "nsash": nsash,
                    "n_sash_active": n_sash_active,
                    "n_sash_passive": n_sash_passive,
                    "hinges_per_sash": hinges_per_sash,
                    "n_rect": n_rect,
                    "n_impost": n_impost,
                }

                total_value += eval_formula(str(formula), ctx)

            gabarit_values.append([type_elem, total_value])

        self.excel.clear_and_write(SHEET_GABARITS, self.HEADER, gabarit_values)

        return gabarit_values, total_area


# ======================================
# РАСЧЕТ МАТЕРИАЛОВ (СПРАВОЧНИК -1)
# ======================================

class MaterialCalculator:
    HEADER = [
        "Тип изделия", "Система профиля", "Тип элемента", "Артикул", "Товар",
        "Ед.", "Цена за ед.", "Ед. фактического расхода",
        "Кол-во фактического расхода (J)",
        "Норма к упаковке", "Ед. к отгрузке",
        "Кол-во к отгрузке", "Сумма"
    ]

    def __init__(self, excel_client: ExcelClient):
        self.excel = excel_client

    def calculate(self, order: dict, positions: list, selected_duplicates: dict):
        ref_rows = self.excel.read_records(SHEET_REF1)
        total_area = sum(p["area_m2"] * p["Nwin"] for p in positions)

        if not ref_rows:
            return [], 0.0, total_area

        try:
            nsash = int(order.get("sashes", "1"))
        except ValueError:
            nsash = 1
        n_sash_active = 1 if nsash >= 1 else 0
        n_sash_passive = max(nsash - 1, 0)
        hinges_per_sash = 3

        result_rows = []
        total_sum = 0.0

        for row in ref_rows:
            row_type = get_field(row, "тип издел", "")
            row_profile = get_field(row, "система проф", "")
            type_elem = get_field(row, "тип элемент", "")
            product_name = str(get_field(row, "товар", "") or "")

            # фильтр по типу изделия
            if row_type:
                if str(row_type).strip().lower() != order["product_type"].strip().lower():
                    continue

            # фильтр по системе профиля
            if row_profile:
                if str(row_profile).strip().lower() != order["profile_system"].strip().lower():
                    continue

            # фильтр по выбору дублей
            if type_elem in selected_duplicates and selected_duplicates[type_elem]:
                chosen_names = selected_duplicates[type_elem]
                if product_name not in chosen_names:
                    continue

            formula = get_field(row, "формула_python", "")
            if not formula:
                formula = get_field(row, "формула фактического расхода", "")

            qty_fact_total = 0.0

            for p in positions:
                width = p["width_mm"]
                height = p["height_mm"]
                left = p["left_mm"]
                center = p["center_mm"]
                right = p["right_mm"]
                top = p["top_mm"]
                sash_w = p["sash_width_mm"]
                sash_h = p["sash_height_mm"]
                area = p["area_m2"]
                perimeter = p["perimeter_m"]
                qty = p["Nwin"]

                n_rect = 1 + (1 if left > 0 else 0) + (1 if top > 0 else 0)
                n_impost = 0

                ctx = {
                    "width": width,
                    "height": height,
                    "left": left,
                    "center": center,
                    "right": right,
                    "top": top,
                    "sash_width": sash_w,
                    "sash_height": sash_h,
                    "area": area,
                    "perimeter": perimeter,
                    "qty": qty,
                    "nsash": nsash,
                    "n_sash_active": n_sash_active,
                    "n_sash_passive": n_sash_passive,
                    "hinges_per_sash": hinges_per_sash,
                    "n_rect": n_rect,
                    "n_impost": n_impost,
                }
                qty_fact_total += eval_formula(str(formula), ctx)

            unit_price = safe_float(get_field(row, "цена за", 0.0))
            norm_per_pack = safe_float(get_field(row, "кол-во норм", 0.0))
            unit_pack = str(get_field(row, "ед .норма к упаковке", "") or "").strip()
            unit = str(get_field(row, "ед.", "") or "").strip()
            unit_fact = str(get_field(row, "ед. фактического расхода", "") or "").strip()

            if norm_per_pack > 0:
                qty_to_ship = math.ceil(qty_fact_total / norm_per_pack)
                effective_qty = qty_to_ship * norm_per_pack
            else:
                qty_to_ship = qty_fact_total
                effective_qty = qty_fact_total

            sum_row = effective_qty * unit_price
            total_sum += sum_row

            result_rows.append([
                row_type if row_type is not None else "",
                row_profile if row_profile is not None else "",
                type_elem,
                get_field(row, "артикул", ""),
                product_name,
                unit,
                unit_price,
                unit_fact,
                qty_fact_total,
                norm_per_pack,
                unit_pack,
                qty_to_ship,
                sum_row
            ])

        self.excel.clear_and_write(SHEET_MATERIAL, self.HEADER, result_rows)

        return result_rows, total_sum, total_area


# ======================================
# РАСЧЕТ ИТОГОВ (СПРАВОЧНИК -2)
# ======================================

class FinalCalculator:
    HEADER = ["Наименование услуг", "Стоимость за м²", "Ед", "Итого"]

    def __init__(self, excel_client: ExcelClient):
        self.excel = excel_client

    def calculate(self, order: dict, total_area: float, material_total: float):
        ref_rows = self.excel.read_records(SHEET_REF2)

        glass_type = order["glass_type"]
        filling = order["filling"]
        toning = order["toning"]
        assembly = order["assembly"]
        montage = order["montage"]

        selected = None
        for row in ref_rows:
            row_type = get_field(row, "тип стеклопак", "")
            row_fill = get_field(row, "заполн", "")
            if (str(row_type).strip() == glass_type and
                    str(row_fill).strip() == filling):
                selected = row
                break

        if not selected and ref_rows:
            selected = ref_rows[0]
        elif not selected:
            selected = {}

        price_glass = safe_float(get_field(selected, "стоимость стеклопак", 0.0))
        # Нарезку больше не используем в итогах:
        # price_cut = safe_float(get_field(selected, "стоимость резки", 0.0))
        price_toning = safe_float(get_field(selected, "стоимость тониров", 0.0))
        price_assembly = safe_float(get_field(selected, "стоимость сборк", 0.0))
        price_montage = safe_float(get_field(selected, "стоимость монтаж", 0.0))

        rows = []

        glass_sum = total_area * price_glass
        rows.append(["Стеклопакет", price_glass, "за м²", glass_sum])

        # Нарезка убрана из расчёта итогов:
        # cut_sum = total_area * price_cut
        # rows.append(["Нарезка", price_cut, "за м²", cut_sum])

        if toning == "Есть":
            toning_sum = total_area * price_toning
            price_toning_use = price_toning
        else:
            toning_sum = 0.0
            price_toning_use = 0.0
        rows.append(["Тонировка", price_toning_use, "за м²", toning_sum])

        if assembly == "Есть":
            assembly_sum = total_area * price_assembly
            price_assembly_use = price_assembly
        else:
            assembly_sum = 0.0
            price_assembly_use = 0.0
        rows.append(["Сборка", price_assembly_use, "за м²", assembly_sum])

        if montage == "Есть":
            montage_sum = total_area * price_montage
            price_montage_use = price_montage
        else:
            montage_sum = 0.0
            price_montage_use = 0.0
        rows.append(["Монтаж", price_montage_use, "за м²", montage_sum])

        rows.append(["Материал", "-", "-", material_total])

        # База без cut_sum
        base_sum = glass_sum + toning_sum + assembly_sum + montage_sum + material_total

        # Обеспечение 60%, а не 70%
        ensure_sum = base_sum * 0.6
        rows.append(["Обеспечение", "", "", ensure_sum])

        total_sum = base_sum + ensure_sum

        extra_rows = [
            ["ИТОГО", "", "", total_sum]
        ]

        self.excel.clear_and_write(SHEET_FINAL, self.HEADER, rows + extra_rows)

        return rows, total_sum, ensure_sum


# ======================================
# ЭКСПОРТ СМЕТЫ В ПАМЯТЬ (ДЛЯ СКАЧИВАНИЯ)
# ======================================

def build_smeta_workbook(order: dict,
                         positions: list,
                         gabarit_rows: list,
                         material_rows: list,
                         final_rows: list,
                         total_area: float,
                         material_total: float,
                         total_sum: float,
                         ensure_sum: float) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    row = 1
    ws.cell(row=row, column=1, value=f"Заказ № {order['order_number']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Тип изделия: {order['product_type']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Вид изделия: {order['product_view']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Створки: {order['sashes']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Профильная система: {order['profile_system']}")
    row += 1
    ws.cell(row=row, column=1,
            value=f"Стеклопакет: {order['glass_thickness']} / {order['glass_type']} / {order['filling']}")
    row += 1
    # Строка без "Нарезка"
    ws.cell(
        row=row,
        column=1,
        value=f"Тонировка: {order['toning']}, Сборка: {order['assembly']}, Монтаж: {order['montage']}"
    )
    row += 1
    ws.cell(row=row, column=1, value=f"Общая площадь (м²): {total_area:.3f}")
    row += 2

    # Позиции
    ws.cell(row=row, column=1, value="Позиции заказа")
    row += 1
    headers_pos = ["№", "Ширина, мм", "Высота, мм", "Nwin", "Площадь, м²", "Периметр, м"]
    for col, h in enumerate(headers_pos, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1

    for i, p in enumerate(positions, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p["width_mm"])
        ws.cell(row=row, column=3, value=p["height_mm"])
        ws.cell(row=row, column=4, value=p["Nwin"])
        ws.cell(row=row, column=5, value=p["area_m2"])
        ws.cell(row=row, column=6, value=p["perimeter_m"])
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Расчет по габаритам")
    row += 1
    ws.cell(row=row, column=1, value="Тип элемента")
    ws.cell(row=row, column=2, value="Фактическое значение")
    row += 1
    for t, v in gabarit_rows:
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=float(v))
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Расчет материалов")
    row += 1
    headers_mat = [
        "Тип изделия", "Система профиля", "Тип элемента", "Артикул", "Товар",
        "Ед.", "Цена за ед.", "Ед. факт. расхода",
        "Кол-во факт. расхода (J)",
        "Норма к упаковке", "Ед. к отгрузке",
        "Кол-во к отгрузке", "Сумма"
    ]
    for col, h in enumerate(headers_mat, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1

    for r in material_rows:
        for col, val in enumerate(r, start=1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Итого по материалам: {material_total:.2f}")
    row += 2

    ws.cell(row=row, column=1, value="Итоговый расчет с монтажом")
    row += 1
    headers_fin = ["Наименование услуг", "Стоимость за м²", "Ед", "Итого"]
    for col, h in enumerate(headers_fin, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1

    for r in final_rows:
        for col, val in enumerate(r, start=1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Обеспечение (60%): {ensure_sum:.2f}")
    row += 1
    ws.cell(row=row, column=1, value=f"ИТОГО к оплате: {total_sum:.2f}")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ======================================
# WEB-ИНТЕРФЕЙС НА STREAMLIT
# ======================================

def main():
    st.set_page_config(page_title="Axis Pro GF • Калькулятор", layout="wide")
    st.title("📘 Калькулятор алюминиевых изделий (Axis Pro GF)")

    st.info(f"Используется файл справочника: **{EXCEL_FILE}**")

    if not os.path.exists(EXCEL_FILE):
        st.error(f"Не найден Excel-файл справочника: {EXCEL_FILE}")
        st.stop()

    excel = ExcelClient(EXCEL_FILE)

    # ---------- Общие данные заказа (в сайдбаре) ----------
    with st.sidebar:
        st.header("Общие данные заказа")

        order_number = st.text_input("Номер заказа", value="")
        product_type = st.selectbox("Тип изделия", ["Окно", "Дверь", "Тамбур"])
        product_view = st.selectbox("Вид изделия", ["Стандарт", "С фрамугой"])
        sashes = st.selectbox("Створки", ["1", "2"])

        profile_system = st.selectbox(
            "Профильная система",
            ["ALG 2030-45C", "ALG RUIT 63i", "ALG RUIT 73", "ALG RUIT 73i", "Другое"]
        )

        glass_thickness = st.selectbox("Стеклопакет (толщина)", ["32 мм", "40 мм", "42 мм", "50 мм"])


        glass_type = st.selectbox(
    "Тип стеклопакета",
    [
        "двойной",
        "тройной",
        "энергодвойной",
        "энерготройной",
        "Одинарный 4мм",
        "Одинарный 6мм",
        "Одинарный 4мм закал",
        "Одинарный 6мм закал",
    ]
)

        filling = st.selectbox("Заполнение", ["СПО", "СПД"])

        # Нарезку из формы убрали
        toning = st.selectbox("Тонировка", ["Нет", "Есть"])
        assembly = st.selectbox("Сборка", ["Нет", "Есть"])
        montage = st.selectbox("Монтаж", ["Нет", "Есть"])

        positions_count = st.number_input("Количество позиций", min_value=1, max_value=10, value=1, step=1)

    # ---------- Позиции ----------
    st.header("🧱 Позиции (габариты изделий)")

    positions_inputs = []
    for i in range(int(positions_count)):
        st.subheader(f"Позиция {i + 1}")
        col1, col2, col3, col4 = st.columns(4)

        width_mm = col1.number_input(f"Ширина, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"w_{i}")
        height_mm = col2.number_input(f"Высота, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"h_{i}")
        left_mm = col3.number_input(f"LEFT, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"l_{i}")
        right_mm = col4.number_input(f"RIGHT, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"r_{i}")

        col5, col6, col7, col8 = st.columns(4)
        center_mm = col5.number_input(f"CENTER, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"c_{i}")
        top_mm = col6.number_input(f"TOP, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"t_{i}")
        sash_width_mm = col7.number_input(f"Ширина створки, мм (поз. {i+1})", min_value=0.0, step=10.0, key=f"sw_{i}")
        sash_height_mm = col8.number_input(f"Высота створки, мм (поз. {i+1})", min_value=0.0, step=10.0,
                                           key=f"sh_{i}")

        col9, _ = st.columns(2)
        Nwin = col9.number_input(f"Кол-во Nwin (поз. {i+1})", min_value=1, step=1, value=1, key=f"nwin_{i}")

        positions_inputs.append({
            "width_mm": width_mm,
            "height_mm": height_mm,
            "left_mm": left_mm,
            "center_mm": center_mm,
            "right_mm": right_mm,
            "top_mm": top_mm,
            "sash_width_mm": sash_width_mm,
            "sash_height_mm": sash_height_mm,
            "Nwin": Nwin,
        })

    # ---------- Выбор материалов при дублях ----------
    st.header("🧾 Выбор материалов при дублях (если в справочнике несколько товаров на один элемент)")
    selected_duplicates = {}

    ref1 = excel.read_records(SHEET_REF1)
    groups = {}
    for row in ref1:
        row_type = str(get_field(row, "тип издел", "") or "").strip()
        row_profile = str(get_field(row, "система проф", "") or "").strip()

        if row_type.lower() != product_type.lower():
            continue
        if row_profile.lower() != profile_system.lower():
            continue

        type_elem = str(get_field(row, "тип элемент", "") or "").strip()
        product_name = str(get_field(row, "товар", "") or "").strip()
        if not type_elem or not product_name:
            continue

        groups.setdefault(type_elem, set()).add(product_name)

    if not groups:
        st.info("Для выбранного типа изделия и профиля дублей материалов не найдено.")
    else:
        for type_elem, products in sorted(groups.items(), key=lambda kv: kv[0]):
            if len(products) <= 1:
                continue
            default = sorted(list(products))
            chosen = st.multiselect(
                f"Тип элемента: {type_elem}",
                options=sorted(list(products)),
                default=default,
                key=f"dup_{type_elem}"
            )
            selected_duplicates[type_elem] = set(chosen)

    # ---------- Кнопка расчёта ----------
    st.markdown("---")
    calc_button = st.button("💾 Сохранить в Excel и выполнить расчёт")

    if calc_button:
        # Проверка заполненности
        if not order_number.strip():
            st.error("Введите номер заказа в левой панели.")
            st.stop()

        positions = []
        for p in positions_inputs:
            if p["width_mm"] <= 0 or p["height_mm"] <= 0:
                st.error("Во всех позициях ширина и высота должны быть больше 0.")
                st.stop()
            area_m2 = (p["width_mm"] * p["height_mm"]) / 1_000_000.0
            perimeter_m = 2 * (p["width_mm"] + p["height_mm"]) / 1000.0

            positions.append({
                **p,
                "area_m2": area_m2,
                "perimeter_m": perimeter_m,
            })

        # Если створки есть, но размеры створки не заданы — принимаем = окну
        try:
            sashes_count = int(sashes)
        except ValueError:
            sashes_count = 1

        if sashes_count >= 1:
            for p in positions:
                if p["sash_width_mm"] <= 0:
                    p["sash_width_mm"] = p["width_mm"]
                if p["sash_height_mm"] <= 0:
                    p["sash_height_mm"] = p["height_mm"]

        order = {
            "order_number": order_number.strip(),
            "product_type": product_type,
            "product_view": product_view,
            "sashes": sashes,
            "profile_system": profile_system,
            "glass_thickness": glass_thickness,
            "glass_type": glass_type,
            "filling": filling,
            # cut убрали из формы, но ключ оставим пустым, чтобы не ломать Excel-структуру
            "cut": "",
            "toning": toning,
            "assembly": assembly,
            "montage": montage,
        }

        # Сохраняем в лист ЗАПРОСЫ
        for idx, p in enumerate(positions, start=1):
            row = [
                order["order_number"],
                idx,
                order["product_type"],
                order["product_view"],
                order["sashes"],
                order["profile_system"],
                order["glass_thickness"],
                order["glass_type"],
                order["filling"],
                p["width_mm"],
                p["height_mm"],
                p["left_mm"],
                p["center_mm"],
                p["right_mm"],
                p["top_mm"],
                p["sash_width_mm"],
                p["sash_height_mm"],
                p["Nwin"],
                order["cut"],      # будет пусто
                order["toning"],
                order["assembly"],
                order["montage"],
            ]
            excel.append_form_row(row)

        # Расчёт по габаритам
        gab_calc = GabaritCalculator(excel)
        gabarit_rows, total_area_gab = gab_calc.calculate(order, positions)

        # Расчёт материалов
        mat_calc = MaterialCalculator(excel)
        material_rows, material_total, total_area_mat = mat_calc.calculate(
            order, positions, selected_duplicates
        )

        total_area = total_area_gab

        # Итоговый расчёт
        final_calc = FinalCalculator(excel)
        final_rows, total_sum, ensure_sum = final_calc.calculate(order, total_area, material_total)

        st.success("Расчёт выполнен. Результаты ниже.")

        tab1, tab2, tab3 = st.tabs(["Габариты", "Материалы", "Итоговый расчет"])

        with tab1:
            st.subheader("Расчет по габаритам")
            if gabarit_rows:
                gab_disp = [
                    {"Тип элемента": t, "Фактическое значение": v}
                    for t, v in gabarit_rows
                ]
                st.dataframe(gab_disp, use_container_width=True)
            st.write(f"Общая площадь: **{total_area:.3f} м²**")

        with tab2:
            st.subheader("Расчёт материалов")
            if material_rows:
                mat_disp = []
                for r in material_rows:
                    mat_disp.append({
                        "Тип изделия": r[0],
                        "Система профиля": r[1],
                        "Тип элемента": r[2],
                        "Артикул": r[3],
                        "Товар": r[4],
                        "Ед.": r[5],
                        "Цена за ед.": round(r[6], 2),
                        "Ед. факт. расхода": r[7],
                        "Кол-во факт. расхода": round(r[8], 3),
                        "Норма к упаковке": r[9],
                        "Ед. к отгрузке": r[10],
                        "Кол-во к отгрузке": round(r[11], 3),
                        "Сумма": round(r[12], 2),
                    })
                st.dataframe(mat_disp, use_container_width=True)
            st.write(f"Итого по материалам: **{material_total:.2f}**")

        with tab3:
            st.subheader("Итоговый расчет с монтажом")
            if final_rows:
                fin_disp = []
                for name, price, unit, total in final_rows:
                    fin_disp.append({
                        "Наименование услуг": name,
                        "Стоимость за м²": price if isinstance(price, str) else round(price, 2),
                        "Ед": unit,
                        "Итого": total if isinstance(total, str) else round(total, 2),
                    })
                st.dataframe(fin_disp, use_container_width=True)
            st.write(f"Обеспечение (60%): **{ensure_sum:.2f}**")
            st.write(f"ИТОГО к оплате: **{total_sum:.2f}**")

        # Кнопка скачивания сметы
        smeta_bytes = build_smeta_workbook(
            order,
            positions,
            gabarit_rows,
            material_rows,
            final_rows,
            total_area,
            material_total,
            total_sum,
            ensure_sum
        )
        default_name = f"Смета_Заказ_{order['order_number']}.xlsx"
        st.download_button(
            "⬇️ Скачать смету в Excel",
            data=smeta_bytes,
            file_name=default_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
