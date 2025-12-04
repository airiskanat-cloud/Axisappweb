import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
from pathlib import Path

# === НАСТРОЙКИ ===
EXCEL_FILE = "axis_pro_gf.xlsx"   # имя файла Excel рядом с этим .py


# === КЛАСС ДЛЯ РАБОТЫ С EXCEL ===
class ExcelClient:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = None
        self.load()

    def load(self):
        """Загружаем книгу Excel."""
        self.wb = load_workbook(self.filename, data_only=True)

    @property
    def sheets(self):
        return self.wb.sheetnames

    def get_sheet_df(self, sheet_name: str) -> pd.DataFrame:
        """Читаем лист в DataFrame (первая строка — заголовки)."""
        ws = self.wb[sheet_name]
        data = list(ws.values)

        if not data:
            return pd.DataFrame()

        header = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=header)
        return df

    def save_df_to_sheet(self, df: pd.DataFrame, sheet_name: str):
        """Полностью перезаписываем лист содержимым df."""
        # если лист есть — удаляем и создаём заново
        if sheet_name in self.wb.sheetnames:
            ws_old = self.wb[sheet_name]
            self.wb.remove(ws_old)
        ws = self.wb.create_sheet(sheet_name)

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        self.wb.save(self.filename)


# === ОСНОВНОЕ ПРИЛОЖЕНИЕ STREAMLIT ===
def main():
    st.set_page_config(page_title="Axis Web", layout="wide")

    st.title("Axis Web – работа с Excel")
    st.write("Файл данных:", f"`{EXCEL_FILE}`")

    # Проверяем, что файл существует
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        st.error(
            f"Файл **{EXCEL_FILE}** не найден в папке проекта.\n"
            "Загрузите его в репозиторий и задеплойте заново."
        )
        st.stop()

    # Загружаем Excel
    try:
        excel = ExcelClient(EXCEL_FILE)
    except BadZipFile:
        st.error(
            f"Файл **{EXCEL_FILE}** повреждён или не является настоящим `.xlsx`.\n"
            "Откройте его в Excel и сохраните как **Книга Excel (*.xlsx)**, "
            "потом замените файл в проекте и сделайте `git add/commit/push`."
        )
        st.stop()
    except Exception as e:
        st.error(f"Ошибка при открытии Excel: {e}")
        st.stop()

    # Выбор листа
    sheet_name = st.sidebar.selectbox("Выберите лист Excel", excel.sheets)

    # Загружаем данные выбранного листа
    df = excel.get_sheet_df(sheet_name)

    st.subheader(f"Данные листа: {sheet_name}")
    if df.empty:
        st.info("На этом листе пока нет данных.")
    else:
        st.dataframe(df, use_container_width=True)

    # Форма добавления строки
    st.subheader("Добавить новую строку")

    if df.empty:
        st.info("Нет колонок для ввода (лист пустой). "
                "Добавьте хотя бы одну строку в Excel вручную.")
        return

    with st.form("add_row_form"):
        inputs = {}
        for col in df.columns:
            # Простейший текстовый ввод для каждого столбца
            val = st.text_input(str(col))
            inputs[col] = val

        submitted = st.form_submit_button("Сохранить строку")

        if submitted:
            new_row = pd.DataFrame([inputs])
            new_df = pd.concat([df, new_row], ignore_index=True)
            try:
                excel.save_df_to_sheet(new_df, sheet_name)
                st.success("Строка сохранена в Excel. Обновите страницу, чтобы увидеть изменения.")
            except Exception as e:
                st.error(f"Ошибка при сохранении в Excel: {e}")


if __name__ == "__main__":
    main()
